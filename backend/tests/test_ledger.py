from pathlib import Path

import aiosqlite
import httpx
import openpyxl
import pytest

from app.core.db import connect
from app.services import ledger, staff

S1 = {"id": "s1", "display_name": "山本", "role": "staff"}


async def _quote(client: httpx.AsyncClient, user: str = "u1") -> dict:
    h = {"X-Test-User": user}
    await client.put("/api/v1/cart/items/DR-100", json={"quantity": 1}, headers=h)
    return (await client.post("/api/v1/quotes", json={}, headers=h)).json()


@pytest.fixture
async def db(db_path: Path) -> aiosqlite.Connection:
    conn = await connect(str(db_path))
    yield conn
    await conn.close()


async def test_export_quotes(
    client: httpx.AsyncClient, db: aiosqlite.Connection, tmp_path: Path
) -> None:
    q1 = await _quote(client, user="u1")
    await _quote(client, user="u2")
    await staff.set_status(db, q1["id"], "ordered")

    path = await ledger.export(db, "quotes", outdir=tmp_path)
    assert path.exists()
    assert path.suffix == ".xlsx"

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "見積番号"  # ヘッダー行
    assert len(rows) == 1 + 2  # ヘッダー+品目2行(1見積1品目)

    quote_nos = {r[0] for r in rows[1:]}
    assert q1["quote_no"] in quote_nos


async def test_export_orders_only_ordered(
    client: httpx.AsyncClient, db: aiosqlite.Connection, tmp_path: Path
) -> None:
    q1 = await _quote(client, user="u1")
    await _quote(client, user="u2")  # requested のまま
    await staff.set_status(db, q1["id"], "ordered")

    path = await ledger.export(db, "orders", outdir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1 + 1  # ヘッダー+受注1件のみ
    assert rows[1][0] == q1["quote_no"]
    assert rows[1][1] == "受注"
