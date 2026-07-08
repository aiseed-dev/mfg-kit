"""台帳 xlsx(S-05)。openpyxl で直接生成し OnlyOffice でそのまま開ける。

帳票の追加は「SELECT+整形の関数を1本足す」だけ(README の原則)。
"""

from datetime import datetime
from pathlib import Path

import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.core.config import settings

# 見積番号・受付/回答/受注日・担当・相手・品目・数量(01_requirements S-05)
COLUMNS = [
    ("見積番号", 12),
    ("状態", 10),
    ("受付日", 12),
    ("回答日", 12),
    ("受注日", 12),
    ("担当", 10),
    ("会社", 18),
    ("担当者", 12),
    ("品目", 28),
    ("型番", 12),
    ("数量", 8),
    ("個別仕様", 24),
]

STATUS_JA = {
    "requested": "依頼中",
    "answered": "回答済み",
    "ordered": "受注",
    "declined": "辞退",
    "cancelled": "取下げ",
    "closed": "完了",
    "expired": "期限切れ",
}

_SQL = """
    SELECT q.quote_no, q.status, q.created_at, q.answered_at, q.ordered_at,
           s.contact_label AS staff_label, u.company_name,
           u.display_name AS customer_name,
           p.name AS product_name, p.code AS product_code,
           qi.quantity, qi.spec_note
    FROM quotes q
    JOIN app_users u ON u.id = q.customer_id
    LEFT JOIN app_users s ON s.id = q.answered_by
    JOIN quote_items qi ON qi.quote_id = q.id
    JOIN products p ON p.id = qi.product_id
    {where}
    ORDER BY q.quote_year, q.quote_seq, p.code
"""


def _day(iso: str | None) -> str:
    return iso[:10] if iso else ""


async def export(
    db: aiosqlite.Connection, kind: str = "quotes", outdir: Path | None = None
) -> Path:
    """kind: quotes=見積台帳(全件) / orders=受注台帳(受注・完了のみ。
    計上基準は ordered_at)。1品目1行。"""
    if kind == "orders":
        where = "WHERE q.status IN ('ordered', 'closed')"
        title = "受注台帳"
    else:
        where = ""
        title = "見積台帳"
    rows = await db.execute_fetchall(_SQL.format(where=where))

    wb = Workbook()
    ws = wb.active
    ws.title = title
    for i, (name, width) in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=i, value=name).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row["quote_no"])
        ws.cell(r, 2, STATUS_JA.get(row["status"], row["status"]))
        ws.cell(r, 3, _day(row["created_at"]))
        ws.cell(r, 4, _day(row["answered_at"]))
        ws.cell(r, 5, _day(row["ordered_at"]))
        ws.cell(r, 6, row["staff_label"])
        ws.cell(r, 7, row["company_name"])
        ws.cell(r, 8, row["customer_name"])
        ws.cell(r, 9, row["product_name"])
        ws.cell(r, 10, row["product_code"])
        ws.cell(r, 11, row["quantity"])
        ws.cell(r, 12, row["spec_note"])
    ws.freeze_panes = "A2"

    outdir = outdir or Path(settings.export_dir)
    outdir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    path = outdir / f"{title}-{stamp}.xlsx"
    wb.save(path)
    return path
